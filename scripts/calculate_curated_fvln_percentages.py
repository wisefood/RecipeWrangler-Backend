#!/usr/bin/env python3
"""Calculate FVLN coverage for the Hungarian and Slovenian curated datasets.

Hungarian recipes use the same cached USDA-ID method as the existing Section 5
reference FVLN calculation. Slovenian recipes use their source-provided English
ingredient names and exact gram weights because their cached EU profile IDs do
not provide adequate USDA food-group mapping coverage.
"""

from __future__ import annotations

import importlib.util
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
OUTPUT = REPO / "section5_outputs/curated_recipe_fvln_percentages.csv"

# Exact source names belonging to the same FVLN groups used by the USDA-prefix
# method: fruit, vegetables, legumes, nuts, and eligible olive/rapeseed oils.
SLOVENIAN_FVLN_INGREDIENTS = {
    "Apple, average",
    "Apple, flesh without skin, raw (average)",
    "Asparagus",
    "BEANS, white, cooked",
    "Banana",
    "Bell pepper,  yellow",
    "Bell pepper, red",
    "Bell pepper, sweet pepper, green",
    "CELERY, Average",
    "CELERY, Bulb",
    "Cabbage, white",
    "Carrot",
    "Cauliflower",
    "Chanterelle or girolle mushroom, raw",
    "Chard",
    "Chicory, red",
    "Chive or spring onion, fresh",
    "Corn-salad",
    "Cucumber",
    "Curly kale, raw",
    "Dandelion",
    "Extra virgin olive oil",
    "GARLIC",
    "Horseradish",
    "Leek",
    "Lemon juice, freshly prepared",
    "Lemon peel, raw",
    "MUSHROOMS",
    "Mint, fresh",
    "Olive, green, in brine",
    "Onion, brown",
    "PARSLEY, Root",
    "PARSLEY, leaves",
    "POTATO, Average",
    "Pea",
    "Peppers, hot chili, green, raw",
    "Potato with coat cooked",
    "Raisins",
    "Rapeseed oil",
    "Spinach",
    "Squash, all types, raw",
    "Strawberries in syrup tinned",
    "TOMATO PUREE",
    "Tomato",
    "Turnip",
    "Walnuts",
    "Zucchini",
    "lemon",
    "sauerkraut",
}


def _load_reference_fvln_module():
    path = REPO / "scripts/calculate_reference_fvln_percentages.py"
    spec = importlib.util.spec_from_file_location("reference_fvln", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Could not load {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def calculate_hungarian() -> list[dict]:
    module = _load_reference_fvln_module()
    canonical, labels = module.load_name_maps()
    profiles = module.psql_csv(
        """
        SELECT source, recipe_id, title,
               nutrition_profiling_details::text AS details
        FROM public."nutrients-recipe-profiles"
        WHERE source = 'Curated Hungarian Recipes'
          AND nutrition_source = 'usda'
          AND nutrition_profiling_details IS NOT NULL
        """
    )
    module.SOURCE_LABELS["Curated Hungarian Recipes"] = (
        "Curated Hungarian Recipes"
    )
    return [
        module.calculate_row(record, canonical, labels)
        for record in profiles.to_dict("records")
    ]


def calculate_slovenian() -> list[dict]:
    workbook = load_workbook(
        REPO / "data/Slovenia/Slovenian_Recipes.xlsx",
        read_only=True,
        data_only=True,
    )
    ingredients = defaultdict(list)
    for row in workbook["Sestavine"].iter_rows(min_row=2, values_only=True):
        recipe_id, name, amount = row[0], row[3], row[5]
        if not recipe_id or not name or amount is None:
            continue
        try:
            weight = float(amount)
        except (TypeError, ValueError):
            continue
        if weight > 0:
            ingredients[str(recipe_id)].append((str(name).strip(), weight))

    recipes = []
    for row in workbook["Recept"].iter_rows(min_row=2, values_only=True):
        recipe_id = str(row[0])
        title = str(row[2] or "")
        items = ingredients[recipe_id]
        total_weight = sum(weight for _, weight in items)
        eligible_weight = sum(
            weight
            for name, weight in items
            if name in SLOVENIAN_FVLN_INGREDIENTS
        )
        recipes.append(
            {
                "reference_source_or_subset": "Curated Slovenian Recipes",
                "recipe_id": recipe_id,
                "title": title,
                "fvln_percent": (
                    eligible_weight / total_weight * 100.0
                    if total_weight
                    else float("nan")
                ),
                "eligible_weight_g": eligible_weight,
                "total_recipe_weight_g": total_weight,
                "classified_weight_coverage_pct": 100.0,
                "calculation_status": "complete",
                "provenance": (
                    "Exact source ingredient names and gram weights classified "
                    "into the USDA-equivalent FVLN groups"
                ),
            }
        )
    workbook.close()
    return recipes


def main() -> None:
    detail = pd.DataFrame(calculate_hungarian() + calculate_slovenian())
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    detail.to_csv(OUTPUT, index=False)
    summary = detail.groupby("reference_source_or_subset").agg(
        recipes=("recipe_id", "size"),
        median_fvln_percent=("fvln_percent", "median"),
        mean_fvln_percent=("fvln_percent", "mean"),
    )
    print(summary.to_string(float_format=lambda value: f"{value:.2f}"))
    print(OUTPUT)


if __name__ == "__main__":
    main()
