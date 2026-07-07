#!/usr/bin/env python3
"""Import Slovenian (OPKP/EuroFIR) recipes into Neo4j, PostgreSQL, and Elasticsearch.

No LLM calls — nutrition comes directly from per-100g EuroFIR values in the workbook.

Usage:
    PYTHONPATH=src .venv/bin/python scripts/import_slovenian.py            # dry run
    PYTHONPATH=src .venv/bin/python scripts/import_slovenian.py --write    # commit
    PYTHONPATH=src .venv/bin/python scripts/import_slovenian.py --write --limit 3
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from recipe_wrangler.utils.env_loader import load_runtime_env

load_runtime_env()

import os

os.environ["LANGCHAIN_TRACING_V2"] = "false"
os.environ["LANGSMITH_TRACING"] = "false"

import openpyxl
import requests

from recipe_wrangler.api.config import get_settings
from recipe_wrangler.repositories.neo4j_recipes import (
    detect_allergens_from_names,
    driver as neo4j_driver,
    upsert_recipe_to_neo4j,
)
from recipe_wrangler.utils.nutri_score import compute_nutri_score_breakdown_from_values
from recipe_wrangler.utils.nutrition_postgres import upsert_recipe_profiling_trace

SOURCE = "Curated Slovenian Recipes"
NUTRITION_SOURCE = "slovenian"
PIPELINE_VERSION = "opkp_direct"
XLSX_FILE = REPO_ROOT / "data" / "Slovenia" / "Slovenian_Recipes.xlsx"
CHECKPOINT_FILE = REPO_ROOT / "scripts" / "import_slovenian.checkpoint.json"

NUTRIENT_MAP = {
    "ENERC": "energy_kcal",
    "PROT": "protein_g",
    "FAT": "fat_g",
    "FASAT": "saturated_fat_g",
    "CHO": "carbohydrate_g",
    "SUGAR": "sugar_g",
    "FIBT": "fibre_g",
    "CA": "calcium_mg",
    "FE": "iron_mg",
    "K": "potassium_mg",
    "FOL": "folate_ug",
    "VITB12": "vitamin_b12_ug",
    "VITC": "vitamin_c_mg",
    "VITD": "vitamin_d_ug",
    "CHORL": "chloride_mg",
}

DISH_TYPE_MAP = {
    "soup": "soup",
    "main dish": "main_dish",
    "salad": "salad",
    "dessert": "dessert",
}


# ---------------------------------------------------------------------------
# Checkpoint
# ---------------------------------------------------------------------------

def load_checkpoint() -> set[str]:
    if CHECKPOINT_FILE.exists():
        return set(json.loads(CHECKPOINT_FILE.read_text()))
    return set()


def save_checkpoint(done: set[str]) -> None:
    CHECKPOINT_FILE.write_text(json.dumps(sorted(done)))


# ---------------------------------------------------------------------------
# Workbook loading -> list of recipe dicts
# ---------------------------------------------------------------------------

def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def load_recipes() -> list[dict]:
    """Parse all three sheets and join them into per-recipe dicts."""
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)

    # Sestavine: recipe_id -> list of {name, amount, unit}
    ingredients_by_recipe: dict[str, list[dict]] = defaultdict(list)
    for row in wb["Sestavine"].iter_rows(min_row=2, values_only=True):
        recipe_id = row[0]
        if not recipe_id:
            continue
        name_en = (row[3] or "").strip()
        ingredients_by_recipe[recipe_id].append(
            {"name": name_en, "amount": _to_float(row[5]), "unit": (row[6] or "g")}
        )

    # Hr. vrednosti: recipe_id -> {field_per_100g}
    nutrition_by_recipe: dict[str, dict] = defaultdict(dict)
    for row in wb["Hr. vrednosti"].iter_rows(min_row=2, values_only=True):
        recipe_id = row[0]
        code = row[1]
        if not recipe_id or not code:
            continue
        value = _to_float(row[2])
        if code == "NACL":
            nutrition_by_recipe[recipe_id]["sodium_mg"] = value * 400.0
        elif code in NUTRIENT_MAP:
            nutrition_by_recipe[recipe_id][NUTRIENT_MAP[code]] = value

    # Recept: the driver sheet (defines the 100 recipes + order)
    recipes: list[dict] = []
    for row in wb["Recept"].iter_rows(min_row=2, values_only=True):
        recipe_id = row[0]
        if not recipe_id:
            continue
        recipes.append(
            {
                "recipe_id": recipe_id,
                "title_sl": (row[1] or "").strip(),
                "title": (row[2] or "").strip(),
                "recamount": _to_float(row[4]),
                "instructions": (row[6] or "").strip(),
                "yield_factor": _to_float(row[7], default=1.0),
                "category_main": (row[8] or "").strip(),
                "total_time": int(_to_float(row[9])),
                "ingredients": ingredients_by_recipe.get(recipe_id, []),
                "per_100g": dict(nutrition_by_recipe.get(recipe_id, {})),
            }
        )
    wb.close()
    return recipes


# ---------------------------------------------------------------------------
# Derived quantities
# ---------------------------------------------------------------------------

def _dish_type(category_main: str) -> str:
    return DISH_TYPE_MAP.get(category_main.strip().lower(), category_main.strip().lower().replace(" ", "_"))


def _compute_serves(ingredients: list[dict], yield_factor: float, recamount: float) -> tuple[int, float]:
    total_weight_g = sum(i["amount"] for i in ingredients) * yield_factor
    serves = round(total_weight_g / recamount) if recamount > 0 else 1
    return max(int(serves), 1), total_weight_g


def _compute_nutri_score(per_100g: dict) -> dict | None:
    if not per_100g.get("energy_kcal"):
        return None
    inputs = {
        "energy": per_100g["energy_kcal"] * 4.184,
        "sugar": per_100g.get("sugar_g", 0),
        "saturated_fats": per_100g.get("saturated_fat_g", 0),
        "sodium": per_100g.get("sodium_mg", 0) / 1000,
        "fibers": per_100g.get("fibre_g", 0),
        "proteins": per_100g.get("protein_g", 0),
        "fruit_percentage": 0,
    }
    breakdown = compute_nutri_score_breakdown_from_values(inputs)
    breakdown["basis"] = "per_100g_source"
    return breakdown


# ---------------------------------------------------------------------------
# Neo4j extra properties
# ---------------------------------------------------------------------------

def _set_slovenian_properties(recipe_id: str, dish_type: str) -> None:
    with neo4j_driver.session() as session:
        session.run(
            """
            MATCH (r:Recipe {recipe_id: $recipe_id})
            SET r.description                   = null,
                r.dish_type                     = $dish_type,
                r.has_slovenian_nutrition       = true,
                r.ground_truth_nutrition_source = 'slovenian',
                r.has_profile                   = true,
                r.language                      = 'en'
            """,
            {"recipe_id": recipe_id, "dish_type": dish_type},
        )


# ---------------------------------------------------------------------------
# Elasticsearch
# ---------------------------------------------------------------------------

def _index_elastic(rec: dict, dish_type: str, allergens: list[str],
                   serves: int, breakdown: dict | None) -> None:
    try:
        settings = get_settings()
        doc = {
            "recipe_id": rec["recipe_id"],
            "title": rec["title"],
            "source": SOURCE,
            "source_id": rec["recipe_id"],
            "serves": serves,
            "duration": rec["total_time"],
            "dish_types": [dish_type],
            "tags": [],
            "allergens": allergens,
            "ingredients": [i["name"] for i in rec["ingredients"]],
            "expert_recipe": True,
            "has_profile": True,
            "has_slovenian_nutrition": True,
            "ground_truth_nutrition_source": NUTRITION_SOURCE,
            "nutri_score_slovenian": breakdown.get("nutri_score") if breakdown else None,
            "nutri_color_slovenian": breakdown.get("color") if breakdown else None,
        }
        requests.put(
            f"{settings.elastic_url}/recipes_v2/_doc/{rec['recipe_id']}",
            json=doc,
            timeout=5,
        ).raise_for_status()
    except Exception as exc:
        print(f"    [ES] WARN {exc}", flush=True)


# ---------------------------------------------------------------------------
# Postgres
# ---------------------------------------------------------------------------

def _upsert_postgres(rec: dict, serves: int, total_weight_g: float,
                     breakdown: dict | None, now_iso: str) -> None:
    per_100g = rec["per_100g"]
    recamount = rec["recamount"]

    per_serving = {field: round(value * recamount / 100.0, 4) for field, value in per_100g.items()}
    total = {field: round(v * serves, 4) for field, v in per_serving.items()}

    nutri_score_jsonb = None
    if breakdown:
        nutri_score_jsonb = {
            "nutri_score": breakdown.get("nutri_score"),
            "color": breakdown.get("color"),
            "score": breakdown.get("score"),
        }

    details = [
        {"name": i["name"], "weight_g": i["amount"], "unit": i["unit"]}
        for i in rec["ingredients"]
    ]

    debug = {
        "method": "opkp_direct",
        "nutrition_basis": "per_100g_source",
        "recamount_g": recamount,
        "yield_factor": rec["yield_factor"],
        "total_recipe_weight_g": total_weight_g,
        "ingredient_count": len(rec["ingredients"]),
        "serves_source": "deterministic_weight_div_recamount",
    }

    upsert_recipe_profiling_trace(
        {
            "recipe_id": rec["recipe_id"],
            "title": rec["title"],
            "source": SOURCE,
            "source_id": rec["recipe_id"],
            "nutrition_source": NUTRITION_SOURCE,
            "total_nutrients": total or None,
            "total_nutrients_per_serving": per_serving or None,
            "nutri_score": nutri_score_jsonb,
            "nutri_score_breakdown": breakdown,
            "nutrition_profiling_details": details or None,
            "nutrition_profiling_debug": debug,
            "trace": {
                "serves": serves,
                "total_weight_g": total_weight_g,
                "title_sl": rec["title_sl"],
            },
            "pipeline_version": PIPELINE_VERSION,
            "computed_at": now_iso,
        }
    )


# ---------------------------------------------------------------------------
# Per-recipe
# ---------------------------------------------------------------------------

def process_recipe(rec: dict, write: bool) -> str:
    recipe_id = rec["recipe_id"]
    ingredients = rec["ingredients"]
    ingredient_names = [i["name"] for i in ingredients]
    measurements = [f"{i['amount']:g}{i['unit']}" for i in ingredients]
    ingredient_lines = [f"{m} {n}" for m, n in zip(measurements, ingredient_names)]
    instructions = [rec["instructions"]] if rec["instructions"] else []

    dish_type = _dish_type(rec["category_main"])
    serves, total_weight_g = _compute_serves(ingredients, rec["yield_factor"], rec["recamount"])
    allergens = detect_allergens_from_names(ingredient_names)
    breakdown = _compute_nutri_score(rec["per_100g"]) if rec["per_100g"] else None

    if not write:
        grade = breakdown.get("nutri_score") if breakdown else None
        return (
            f"DRY recipe_id={recipe_id} serves={serves} dish={dish_type} "
            f"ingredients={len(ingredient_names)} allergens={allergens} nutri_score={grade}"
        )

    now_iso = datetime.now(timezone.utc).isoformat()

    upsert_recipe_to_neo4j(
        recipe_id=recipe_id,
        title=rec["title"],
        ingredient_lines=ingredient_lines,
        ingredient_names=ingredient_names,
        measurements=measurements,
        instructions=instructions,
        duration=float(rec["total_time"]),
        serves=float(serves),
        image_url=None,
        allergens=allergens,
        tags=[],
        source=SOURCE,
        source_id=recipe_id,
        expert_recipe=True,
    )
    _set_slovenian_properties(recipe_id, dish_type)
    _upsert_postgres(rec, serves, total_weight_g, breakdown, now_iso)
    _index_elastic(rec, dish_type, allergens, serves, breakdown)

    grade = breakdown.get("nutri_score") if breakdown else "n/a"
    return f"WROTE serves={serves} dish={dish_type} ingredients={len(ingredient_names)} grade={grade}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--write", action="store_true", help="Commit to Neo4j/Postgres/ES")
    ap.add_argument("--limit", type=int, default=None, help="Cap recipes for smoke test")
    ap.add_argument("--no-resume", action="store_true", help="Ignore checkpoint")
    args = ap.parse_args()

    recipes = load_recipes()
    done = set() if args.no_resume else load_checkpoint()

    pending = [r for r in recipes if r["recipe_id"] not in done]
    if args.limit:
        pending = pending[: args.limit]

    print(
        f"Slovenian import — {len(pending)} pending / {len(done)} already done / write={args.write}"
    )

    imported = failed = 0
    for i, rec in enumerate(pending):
        recipe_id = rec["recipe_id"]
        print(f"[{i+1}/{len(pending)}] {recipe_id} — {rec['title'][:50]}", end=" ", flush=True)
        try:
            msg = process_recipe(rec, write=args.write)
            print(msg, flush=True)
            if args.write:
                done.add(recipe_id)
                save_checkpoint(done)
            imported += 1
        except Exception as exc:
            print(f"ERROR {exc}", flush=True)
            failed += 1

    print(f"\nDone — imported={imported} failed={failed}")


if __name__ == "__main__":
    main()
