#!/usr/bin/env python3
"""Run regional nutrition profiling for Slovenian recipes.

For each recipe x region (irish, hungarian, eu) this script:
  1. calls nutritional_tool_chroma (Chroma match -> Postgres per-100g -> scale -> aggregate),
  2. computes a per-100g Nutri-Score,
  3. upserts a profiling trace row into Postgres
     (source="Curated Slovenian Recipes", nutrition_source=region,
      pipeline_version="opkp_direct_weight_known"),
  4. patches the recipe's recipes_v2 Elasticsearch doc with the per-region
     Nutri-Score grade and color.

Dry-run by default. Pass --write to enable Postgres + Elasticsearch writes.

Usage:
    PYTHONPATH=src .venv/bin/python scripts/profile_slovenian_regions.py
    PYTHONPATH=src .venv/bin/python scripts/profile_slovenian_regions.py --write
    PYTHONPATH=src .venv/bin/python scripts/profile_slovenian_regions.py --write --limit 2 --regions irish
"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import requests

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "src"))

from recipe_wrangler.utils.env_loader import load_runtime_env

load_runtime_env()

import openpyxl

from recipe_wrangler.api.main import get_settings  # noqa: E402
from recipe_wrangler.tools.nutritional_calculator import nutritional_tool_chroma  # noqa: E402
from recipe_wrangler.utils.nutri_score import (  # noqa: E402
    compute_nutri_score_breakdown_from_values,
)
from recipe_wrangler.utils.nutrition_postgres import (  # noqa: E402
    upsert_recipe_profiling_trace,
)

SOURCE = "Curated Slovenian Recipes"
PIPELINE_VERSION = "opkp_direct_weight_known"
REGIONS = ["irish", "hungarian", "eu"]
ES_REGION_CODE = {"irish": "ie", "hungarian": "hu", "eu": "eu"}

XLSX_FILE = REPO_ROOT / "data" / "Slovenia" / "Slovenian_Recipes.xlsx"
CHECKPOINT_PATH = REPO_ROOT / "scripts" / "profile_slovenian_regions.checkpoint.json"


# ---------------------------------------------------------------------------
# Load recipes from the workbook (same join as the importer)
# ---------------------------------------------------------------------------

def _to_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _load_recipes() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True, data_only=True)

    ingredients_by_recipe: dict[str, list[dict]] = defaultdict(list)
    for row in wb["Sestavine"].iter_rows(min_row=2, values_only=True):
        recipe_id = row[0]
        if not recipe_id:
            continue
        ingredients_by_recipe[recipe_id].append(
            {"name": (row[3] or "").strip(), "amount": _to_float(row[5])}
        )

    recipes: list[dict] = []
    for row in wb["Recept"].iter_rows(min_row=2, values_only=True):
        recipe_id = row[0]
        if not recipe_id:
            continue
        recipes.append(
            {
                "recipe_id": recipe_id,
                "title": (row[2] or "").strip(),
                "recamount": _to_float(row[4]),
                "yield_factor": _to_float(row[7], default=1.0),
                "ingredients": ingredients_by_recipe.get(recipe_id, []),
            }
        )
    wb.close()
    return recipes


# ---------------------------------------------------------------------------
# Checkpoint (atomic, key = recipe_id:region)
# ---------------------------------------------------------------------------

def _load_checkpoint(resume: bool) -> set[str]:
    if not resume or not CHECKPOINT_PATH.exists():
        return set()
    try:
        import json
        return set(json.loads(CHECKPOINT_PATH.read_text()))
    except Exception:
        return set()


def _save_checkpoint(done: set[str]) -> None:
    import json
    tmp = CHECKPOINT_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(sorted(done), indent=2))
    os.replace(tmp, CHECKPOINT_PATH)


# ---------------------------------------------------------------------------
# Nutri-Score per 100g (from Chroma totals + total weight)
# ---------------------------------------------------------------------------

def _compute_nutri_score_per100g(totals: dict, total_weight_g: float) -> dict | None:
    if total_weight_g <= 0:
        return None
    scale = 100.0 / total_weight_g
    inputs = {
        "energy": totals["energy_kcal"] * scale * 4.184,
        "sugar": totals["sugar_g"] * scale,
        "saturated_fats": totals["saturated_fat_g"] * scale,
        "sodium": totals["sodium_mg"] * scale / 1000,
        "fibers": totals["fibre_g"] * scale,
        "proteins": totals["protein_g"] * scale,
        "fruit_percentage": 0,
    }
    breakdown = compute_nutri_score_breakdown_from_values(inputs)
    breakdown["basis"] = "per_100g_from_weight"
    return breakdown


def _patch_elasticsearch(settings, recipe_id: str, region: str, breakdown: dict) -> None:
    code = ES_REGION_CODE[region]
    requests.post(
        f"{settings.elastic_url}/recipes_v2/_update/{recipe_id}",
        json={
            "doc": {
                f"nutri_score_{code}": breakdown.get("nutri_score"),
                f"nutri_color_{code}": breakdown.get("color"),
            }
        },
        timeout=5,
    ).raise_for_status()


# ---------------------------------------------------------------------------
# One recipe x one region
# ---------------------------------------------------------------------------

def _process_region(rec: dict, region: str, settings, write: bool) -> str:
    recipe_id = rec["recipe_id"]
    title = rec["title"]
    ingredients = rec["ingredients"]
    if not ingredients:
        return "skip-no-ingredients"

    ingredients = [i for i in ingredients if i.get("amount") is not None and float(i.get("amount", 0)) > 0]
    if not ingredients:
        return "skip-no-ingredients"
    names = [i["name"] for i in ingredients]
    weights = [float(i["amount"]) for i in ingredients]
    total_weight_g = float(sum(weights)) * rec["yield_factor"]

    serves = round(total_weight_g / rec["recamount"]) if rec["recamount"] > 0 else 1
    serves = max(int(serves), 1)

    result = nutritional_tool_chroma.invoke(
        {
            "title": title,
            "ingredient_names": names,
            "weights": weights,
            "source": region,
            "serves": serves,
        }
    )
    totals = result["clean_totals"]
    per_serving_totals = {k: (float(v) / serves if serves else None) for k, v in totals.items()}

    breakdown = _compute_nutri_score_per100g(totals, total_weight_g)

    if not write:
        grade = breakdown.get("nutri_score") if breakdown else None
        return f"dry-run grade={grade}"

    upsert_recipe_profiling_trace(
        {
            "recipe_id": recipe_id,
            "title": title,
            "source": SOURCE,
            "source_id": recipe_id,
            "nutrition_source": region,
            "pipeline_version": PIPELINE_VERSION,
            "total_nutrients": totals,
            "total_nutrients_per_serving": per_serving_totals,
            "nutri_score": {
                "nutri_score": breakdown.get("nutri_score"),
                "color": breakdown.get("color"),
                "score": breakdown.get("score"),
            }
            if breakdown
            else None,
            "nutri_score_breakdown": breakdown,
            "nutrition_profiling_details": result["details"],
            "nutrition_profiling_debug": {
                "method": "opkp_direct_weight_known",
                "region": region,
                "ingredient_count": len(names),
                "total_weight_g": total_weight_g,
            },
            "trace": {
                "serves": serves,
                "serves_source": "deterministic_weight_div_recamount",
                "total_weight_g": total_weight_g,
            },
            "computed_at": datetime.now(timezone.utc).isoformat(),
        }
    )

    if breakdown:
        _patch_elasticsearch(settings, recipe_id, region, breakdown)

    grade = breakdown.get("nutri_score") if breakdown else None
    return f"written grade={grade}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Regional profiling for Slovenian recipes.")
    parser.add_argument("--write", action="store_true", help="enable DB + ES writes (default: dry-run)")
    parser.add_argument("--no-resume", action="store_true", help="ignore checkpoint, re-process all")
    parser.add_argument("--limit", type=int, default=None, help="stop after N recipes")
    parser.add_argument("--regions", type=str, default=None, help="comma-separated subset, e.g. irish,eu")
    args = parser.parse_args()

    regions = REGIONS
    if args.regions:
        requested = [r.strip() for r in args.regions.split(",") if r.strip()]
        unknown = [r for r in requested if r not in REGIONS]
        if unknown:
            parser.error(f"unknown region(s): {unknown}; valid: {REGIONS}")
        regions = [r for r in REGIONS if r in requested]

    settings = get_settings()
    recipes = _load_recipes()
    if args.limit is not None:
        recipes = recipes[: args.limit]

    done = _load_checkpoint(resume=not args.no_resume)
    mode = "WRITE" if args.write else "DRY-RUN"
    print(f"[{mode}] {len(recipes)} recipes x {regions} | resume={len(done)} done")

    processed = 0
    for rec in recipes:
        recipe_id = rec["recipe_id"]
        for region in regions:
            key = f"{recipe_id}:{region}"
            if key in done:
                continue
            try:
                status = _process_region(rec, region, settings, args.write)
                print(f"  {key}: {status}")
            except Exception as exc:
                print(f"  {key}: ERROR {type(exc).__name__}: {exc}")
                continue
            if args.write and status != "skip-no-ingredients":
                done.add(key)
                _save_checkpoint(done)
            processed += 1

    print(f"Done. processed={processed} write={args.write}")


if __name__ == "__main__":
    main()
