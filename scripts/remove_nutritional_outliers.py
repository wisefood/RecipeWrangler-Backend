import json
import re
from pathlib import Path
import openpyxl

# Thresholds from analysis
THRESHOLDS = {
    "HealthyFoods": 806.50,
    "Recipe1M": 29122.33,
    "Irish_SafeFood": 777.75,
    "Slovenian_OPKP": 640.88,
    "MyPlate": 861.28
}

def parse_calories(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.search(r"([\d.]+)", str(val))
    if m:
        return float(m.group(1))
    return None

def clean_healthy_foods(removed_list):
    path = Path("data/HealthyFoods/HealthyFood_recipes_nutrition.json")
    out_path = Path("data/HealthyFoods/HealthyFood_recipes_nutrition_clean.json")
    if not path.exists():
        return
    
    with open(path, "r") as f:
        data = json.load(f)
    
    threshold = THRESHOLDS["HealthyFoods"]
    cleaned_recipes = []
    for r in data["recipes"]:
        c = parse_calories(r.get("nutrition_per_serve", {}).get("Calories"))
        if c is None or c <= threshold:
            cleaned_recipes.append(r)
        else:
            removed_list.append({"dataset": "HealthyFoods", "title": r.get("title", "Unknown"), "calories": c})
            
    data["recipes"] = cleaned_recipes
    data["count"] = len(cleaned_recipes)
    
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)

def clean_recipe1m(removed_list):
    path = Path("data/processed/recipe1m/recipes_with_nutritional_info.json")
    out_path = Path("data/processed/recipe1m/recipes_with_nutritional_info_clean.json")
    if not path.exists():
        return
    
    with open(path, "r") as f:
        data = json.load(f)
    
    threshold = THRESHOLDS["Recipe1M"]
    cleaned_data = []
    for r in data:
        total_nrg = sum(parse_calories(ing.get("nrg")) or 0 for ing in r.get("nutr_per_ingredient", []))
        if total_nrg <= threshold:
            cleaned_data.append(r)
        else:
            removed_list.append({"dataset": "Recipe1M", "title": r.get("title", "Unknown"), "calories": total_nrg})
            
    with open(out_path, "w") as f:
        json.dump(cleaned_data, f, indent=4)

def clean_irish_safefood(removed_list):
    path = Path("data/mappings/Irish Recipes_SafeFood.xlsx")
    out_path = Path("data/mappings/Irish Recipes_SafeFood_Clean.xlsx")
    if not path.exists():
        return
    
    wb = openpyxl.load_workbook(path)
    ws = wb["in"]
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    
    threshold = THRESHOLDS["Irish_SafeFood"]
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "in"
    
    for i, row in enumerate(rows):
        if i < 2: 
            new_ws.append([cell.value for cell in row])
            continue
        
        val = parse_calories(row[25].value)
        if val is None or val <= threshold:
            new_ws.append([cell.value for cell in row])
        else:
            removed_list.append({"dataset": "Irish_SafeFood", "title": row[1].value, "calories": val})
            
    new_wb.save(out_path)

def clean_slovenian_opkp(removed_list):
    path = Path("data/Slovenian_OPKP/primeri_receptov_all_info_per_recipe.json")
    out_path = Path("data/Slovenian_OPKP/primeri_receptov_all_info_per_recipe_clean.json")
    if not path.exists():
        return
    
    with open(path, "r") as f:
        data = json.load(f)
    
    threshold = THRESHOLDS["Slovenian_OPKP"]
    cleaned_recipes = {}
    for r_id, r_data in data["recipes"].items():
        val_to_check = 0
        for nv in r_data.get("nutritional_values", []):
            if nv.get("EuroFIR component code [ecompid] (components thesauri)") == "ENERC":
                val_to_check = parse_calories(nv.get("selected value [selval]"))
                unit = nv.get("unit  (units thesauri) [unit]")
                if unit == "kJ":
                    val_to_check = val_to_check / 4.184
                break
        
        if val_to_check <= threshold:
            cleaned_recipes[r_id] = r_data
        else:
            removed_list.append({"dataset": "Slovenian_OPKP", "title": r_data["recipe"].get("English food name [engfdnam]", "Unknown"), "calories": val_to_check})
            
    data["recipes"] = cleaned_recipes
    data["recipes_count"] = len(cleaned_recipes)
    
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)

def clean_myplate(removed_list):
    path = Path("data/MyPlate/myplate_recipes_nutrition_usda.json")
    out_path = Path("data/MyPlate/myplate_recipes_nutrition_usda_clean.json")
    if not path.exists():
        return
    
    with open(path, "r") as f:
        data = json.load(f)
    
    threshold = THRESHOLDS["MyPlate"]
    cleaned_data = []
    for r in data:
        c = r.get("totals_per_serving_usda", {}).get("energy_kcal")
        c = parse_calories(c)
        if c is None or c <= threshold:
            cleaned_data.append(r)
        else:
            removed_list.append({"dataset": "MyPlate", "title": r.get("title", "Unknown"), "calories": c})
            
    with open(out_path, "w") as f:
        json.dump(cleaned_data, f, indent=2)

def main():
    removed_list = []
    print("Cleaning datasets...")
    clean_healthy_foods(removed_list)
    clean_recipe1m(removed_list)
    clean_irish_safefood(removed_list)
    clean_slovenian_opkp(removed_list)
    clean_myplate(removed_list)
    
    print(f"Total removed: {len(removed_list)}")
    
    # Generate Report
    report_md = "# Detailed Removal Report: Nutritional Outliers\n\n"
    report_md += f"Total recipes removed across all datasets: **{len(removed_list)}**\n\n"
    
    datasets = sorted(list(set(r["dataset"] for r in removed_list)))
    for ds in datasets:
        ds_removed = [r for r in removed_list if r["dataset"] == ds]
        report_md += f"## {ds} ({len(ds_removed)} removed)\n\n"
        report_md += "| Title | Calories | Status |\n"
        report_md += "|---|---|---|\n"
        # Sort by calories descending, show top 50 if too many
        ds_removed.sort(key=lambda x: x["calories"], reverse=True)
        for r in ds_removed[:50]:
            report_md += f"| {r['title']} | {r['calories']:.2f} | Removed |\n"
        if len(ds_removed) > 50:
            report_md += f"| ... and {len(ds_removed) - 50} more | | |\n"
            
    with open("removed_outliers_detailed_report.md", "w") as f:
        f.write(report_md)
    
    print("Removal complete. Detailed report generated in 'removed_outliers_detailed_report.md'.")

if __name__ == "__main__":
    main()
