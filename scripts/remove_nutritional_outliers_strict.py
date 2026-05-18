import json
import re
from pathlib import Path
import openpyxl
from neo4j import GraphDatabase

# NEW Threshold requested by user: 1000 kcal per serving
THRESHOLD_PER_SERVING = 1000.0

def parse_calories(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.search(r"([\d.]+)", str(val))
    if m:
        return float(m.group(1))
    return None

def parse_servings(val):
    if val is None:
        return 1.0
    if isinstance(val, (int, float)):
        return float(val)
    m = re.findall(r"(\d+)", str(val))
    if m:
        vals = [float(x) for x in m]
        return sum(vals) / len(vals)
    return 1.0

def get_recipe1m_servings():
    servings_map = {}
    try:
        driver = GraphDatabase.driver("bolt://localhost:7687", auth=("neo4j", "password123"))
        with driver.session() as session:
            result = session.run("MATCH (r:Recipe) WHERE r.serves IS NOT NULL RETURN r.id, r.serves")
            for record in result:
                servings_map[record["r.id"]] = parse_servings(record["r.serves"])
    except Exception as e:
        print(f"Could not connect to Neo4j: {e}")
    return servings_map

def clean_healthy_foods(removed_list):
    path = Path("data/HealthyFoods/HealthyFood_recipes_nutrition.json")
    out_path = Path("data/HealthyFoods/HealthyFood_recipes_nutrition_clean.json")
    if not path.exists(): return
    with open(path, "r") as f: data = json.load(f)
    cleaned = []
    for r in data["recipes"]:
        c = parse_calories(r.get("nutrition_per_serve", {}).get("Calories"))
        if c is None or c <= THRESHOLD_PER_SERVING: cleaned.append(r)
        else: removed_list.append({"dataset": "HealthyFoods", "title": r.get("title", "Unknown"), "calories": c})
    data["recipes"] = cleaned
    data["count"] = len(cleaned)
    with open(out_path, "w") as f: json.dump(data, f, indent=2)

def clean_recipe1m(removed_list, serv_map):
    path = Path("data/processed/recipe1m/recipes_with_nutritional_info.json")
    out_path = Path("data/processed/recipe1m/recipes_with_nutritional_info_clean.json")
    if not path.exists(): return
    with open(path, "r") as f: data = json.load(f)
    cleaned = []
    for r in data:
        total_nrg = sum(parse_calories(ing.get("nrg")) or 0 for ing in r.get("nutr_per_ingredient", []))
        serves = serv_map.get(r["id"], 4.0)
        per_serving = total_nrg / serves
        if per_serving <= THRESHOLD_PER_SERVING: cleaned.append(r)
        else: removed_list.append({"dataset": "Recipe1M", "title": r.get("title", "Unknown"), "calories": per_serving})
    with open(out_path, "w") as f: json.dump(cleaned, f, indent=4)

def clean_irish_safefood(removed_list):
    path = Path("data/mappings/Irish Recipes_SafeFood.xlsx")
    out_path = Path("data/mappings/Irish Recipes_SafeFood_Clean.xlsx")
    if not path.exists(): return
    wb = openpyxl.load_workbook(path)
    ws = wb["in"]
    rows = list(ws.iter_rows(min_row=1, values_only=False))
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = "in"
    for i, row in enumerate(rows):
        if i < 2: 
            new_ws.append([cell.value for cell in row])
            continue
        val = parse_calories(row[25].value)
        if val is None or val <= THRESHOLD_PER_SERVING: new_ws.append([cell.value for cell in row])
        else: removed_list.append({"dataset": "Irish_SafeFood", "title": row[1].value, "calories": val})
    new_wb.save(out_path)

def clean_myplate(removed_list):
    path = Path("data/MyPlate/myplate_recipes_nutrition_usda.json")
    out_path = Path("data/MyPlate/myplate_recipes_nutrition_usda_clean.json")
    if not path.exists(): return
    with open(path, "r") as f: data = json.load(f)
    cleaned = []
    for r in data:
        c = parse_calories(r.get("totals_per_serving_usda", {}).get("energy_kcal"))
        if c is None or c <= THRESHOLD_PER_SERVING: cleaned.append(r)
        else: removed_list.append({"dataset": "MyPlate", "title": r.get("title", "Unknown"), "calories": c})
    with open(out_path, "w") as f: json.dump(cleaned, f, indent=2)

def main():
    removed_list = []
    print(f"Cleaning datasets using STRICT threshold: {THRESHOLD_PER_SERVING} kcal/serving...")
    serv_map = get_recipe1m_servings()
    clean_healthy_foods(removed_list)
    clean_recipe1m(removed_list, serv_map)
    clean_irish_safefood(removed_list)
    clean_myplate(removed_list)
    
    print(f"Total removed: {len(removed_list)}")
    
    report_md = "# Strict Removal Report: Nutritional Outliers (>1000 kcal/serving)\n\n"
    report_md += f"Total recipes removed across all datasets: **{len(removed_list)}**\n\n"
    
    datasets = sorted(list(set(r["dataset"] for r in removed_list)))
    for ds in datasets:
        ds_removed = [r for r in removed_list if r["dataset"] == ds]
        report_md += f"## {ds} ({len(ds_removed)} removed)\n\n"
        report_md += "| Title | kcal/serving | Status |\n"
        report_md += "|---|---|---|\n"
        ds_removed.sort(key=lambda x: x["calories"], reverse=True)
        for r in ds_removed[:100]:
            report_md += f"| {r['title']} | {r['calories']:.2f} | Removed |\n"
        if len(ds_removed) > 100:
            report_md += f"| ... and {len(ds_removed) - 100} more | | |\n"
            
    with open("removed_outliers_strict_report.md", "w") as f:
        f.write(report_md)
    print("Removal complete. Report generated in 'removed_outliers_strict_report.md'.")

if __name__ == "__main__":
    main()
