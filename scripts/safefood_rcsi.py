"""Helpers for merging RCSI SafeFood lab nutrition onto web SafeFood recipes."""

from __future__ import annotations

import difflib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import NAMESPACE_URL, uuid5

import openpyxl

SOURCE = "Curated Irish Recipes"
LAB_NUTRITION_SOURCE = "safefood_rcsi"
WEB_NUTRITION_SOURCE = "safefood_web"

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_LAB_XLSX = (
    REPO_ROOT / "data" / "Curated Irish Recipes" / "Irish Recipes_updated_22.6.26.xlsx"
)
DEFAULT_WEB_EXPORTS = tuple(
    REPO_ROOT / "exports" / f"safefood_{category}_recipes.json"
    for category in ("breakfast", "lunch", "dinner", "snacks", "desserts")
)

_PREFIX_RE = re.compile(
    r"^(?:AP|DBOT|EmcC|RB|SG|Secondary Analysis)[_\s-]*"
    r"(?:(?:Pimary|Primary)\s+Analysis|primaryanalysis|primary\s+analysis)?[_\s-]*",
    re.I,
)
_DATE_SUFFIX_RE = re.compile(r"[_\s.]*\d{2}\.\d{2}\.\d{2,4}$")

_MANUAL_TITLE_ALIASES = {
    "bang bang chicken": "bang bang chicken salad",
    "sweet chili chicken stuffed peppers": "sweet chilli chicken stuffed peppers",
    "blueberry muffins": "quick and easy blueberry muffins",
    "codandcouscous": "cod and couscous",
    "chickenfingers": "chicken fingers",
    "mixedbeansalad": "mixed bean salad",
    "lentilandvegetablestew": "lentil and vegetable stew",
    "pea soup and bacon": "pea soup with bacon",
    "grilled salmon and warm potato salad": "grilled salmon and potato salad",
    "chickpea soup with red pepper": "chickpea soup with red peppers",
    "baked oat squares": "baked oat bars",
    "cod traybake": "cod tray bake",
    "secondary analysis turkey noodle soup": "turkey noodle soup",
    "chicken fajita": "chicken fajitas",
    "chicken on a stick": "chicken skewers",
    "chicken and broccoli": "chicken and broccoli bake",
    "homemade fruity nutty muesli": "homemade muesli with fruit and nuts",
}


@dataclass(frozen=True)
class RcsiLabRecipe:
    recipe_id_src: str
    raw_title: str
    title: str
    normalized_title: str
    serves: float
    serving_weight_g: float | None
    duration_minutes: float
    cost_category: str | None
    ingredient_lines: list[str]
    instructions: list[str]
    ground_truth_per_serving: dict[str, float]
    ground_truth_per_100g: dict[str, float]


@dataclass(frozen=True)
class SafefoodWebRecipe:
    recipe_id: str
    title: str
    normalized_title: str
    url: str | None
    category: str | None
    raw: dict[str, Any]


def web_recipe_id(title: str) -> str:
    return str(uuid5(NAMESPACE_URL, f"{SOURCE}:{title.strip()}"))


def clean_lab_title(raw: object) -> str:
    title = str(raw or "").strip()
    title = _PREFIX_RE.sub("", title)
    title = _DATE_SUFFIX_RE.sub("", title)
    return title.strip(" _-.")


def normalize_title(raw: object) -> str:
    title = clean_lab_title(raw).lower().replace("&", "and")
    title = title.replace("chili", "chilli")
    title = re.sub(r"\bprimaryanalysis\b|\bprimary analysis\b", " ", title)
    title = re.sub(r"[^a-z0-9]+", " ", title)
    normalized = re.sub(r"\s+", " ", title).strip()
    return _MANUAL_TITLE_ALIASES.get(normalized, normalized)


def parse_minutes(value: object) -> float:
    if not value:
        return 0.0
    text = str(value).lower()
    hours = re.search(r"(\d+(?:\.\d+)?)\s*(?:hr|hour)", text)
    mins = re.search(r"(\d+(?:\.\d+)?)\s*min", text)
    total = 0.0
    if hours:
        total += float(hours.group(1)) * 60.0
    if mins:
        total += float(mins.group(1))
    return total


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None


def _split_lines(value: object) -> list[str]:
    return [line.strip() for line in str(value or "").splitlines() if line.strip()]


def _nutrient(value: object) -> float:
    parsed = parse_number(value)
    return float(parsed or 0.0)


def normalize_cost_category(value: object) -> str | None:
    category = str(value or "").strip().lower()
    return category if category in {"low", "medium", "high"} else None


def _normalized_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _recipe_sheet_rows(
    workbook: openpyxl.Workbook,
) -> list[tuple[dict[str, int], tuple[object, ...]]]:
    """Return rows from every worksheet containing the RCSI recipe schema."""
    rows: list[tuple[dict[str, int], tuple[object, ...]]] = []
    for worksheet in workbook.worksheets:
        header_row: int | None = None
        headers: dict[str, int] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
            start=1,
        ):
            candidate = {
                _normalized_header(value): index
                for index, value in enumerate(row)
                if value is not None
            }
            if "recipe id" in candidate and "recipe name" in candidate:
                header_row = row_number
                headers = candidate
                break
        if header_row is None:
            continue
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            title_index = headers["recipe name"]
            if len(row) > title_index and row[title_index]:
                rows.append((headers, row))
    return rows


def _row_value(
    headers: dict[str, int],
    row: tuple[object, ...],
    name: str,
) -> object:
    index = headers.get(_normalized_header(name))
    return row[index] if index is not None and index < len(row) else None


def _safefood_ground_truth(
    *,
    serves: float,
    energy_kcal_srv: object,
    fat_srv: object,
    satfat_srv: object,
    carb_srv: object,
    sugars_srv: object,
    fibre_srv: object,
    protein_srv: object,
    salt_srv: object,
    energy_kcal_100g: object,
    fat_100g: object,
    satfat_100g: object,
    carb_100g: object,
    sugars_100g: object,
    fibre_100g: object,
    protein_100g: object,
    salt_100g: object,
) -> tuple[dict[str, float], dict[str, float], dict[str, float]]:
    per_serving = {
        "energy_kcal": _nutrient(energy_kcal_srv),
        "fat_g": _nutrient(fat_srv),
        "saturated_fat_g": _nutrient(satfat_srv),
        "carbohydrate_g": _nutrient(carb_srv),
        "sugar_g": _nutrient(sugars_srv),
        "fibre_g": _nutrient(fibre_srv),
        "protein_g": _nutrient(protein_srv),
        "sodium_mg": round(_nutrient(salt_srv) * 400.0, 2),
    }
    per_100g = {
        "energy_kcal": _nutrient(energy_kcal_100g),
        "fat_g": _nutrient(fat_100g),
        "saturated_fat_g": _nutrient(satfat_100g),
        "carbohydrate_g": _nutrient(carb_100g),
        "sugar_g": _nutrient(sugars_100g),
        "fibre_g": _nutrient(fibre_100g),
        "protein_g": _nutrient(protein_100g),
        "sodium_mg": round(_nutrient(salt_100g) * 400.0, 2),
    }
    totals = {key: round(value * serves, 6) for key, value in per_serving.items()}
    return totals, per_serving, per_100g


def load_rcsi_lab_recipes(path: Path = DEFAULT_LAB_XLSX) -> list[RcsiLabRecipe]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    recipes: list[RcsiLabRecipe] = []
    seen_recipe_ids: set[str] = set()
    for headers, row in _recipe_sheet_rows(workbook):
        recipe_id_src = _row_value(headers, row, "Recipe ID")
        recipe_id_key = str(recipe_id_src or "").strip()
        if recipe_id_key and recipe_id_key in seen_recipe_ids:
            continue
        if recipe_id_key:
            seen_recipe_ids.add(recipe_id_key)

        raw_title = _row_value(headers, row, "Recipe Name")
        servings = _row_value(headers, row, "Recipe Servings")
        prep_time = _row_value(headers, row, "Prep Time")
        cook_time = _row_value(headers, row, "Cooking Time")
        cost_category = normalize_cost_category(_row_value(headers, row, "Cost"))
        ingredients_raw = _row_value(headers, row, "Ingredients")
        instructions_raw = _row_value(headers, row, "Cooking instructions")
        portion_measure = _row_value(headers, row, "Portion Measure (g)")
        energy_kcal_100g = _row_value(headers, row, "Energy Kcal (kcal) per 100g")
        fat_100g = _row_value(headers, row, "Fat (g) per 100g")
        satfat_100g = _row_value(headers, row, "Saturated Fat (g) per 100g")
        carb_100g = _row_value(headers, row, "Carbohydrate (g) per 100g")
        sugars_100g = _row_value(headers, row, "Sugars (g) per 100g")
        fibre_100g = _row_value(headers, row, "Fibre (g) per 100g")
        protein_100g = _row_value(headers, row, "Protein (g) per 100g")
        salt_100g = _row_value(headers, row, "Salt (g) per 100g")
        energy_kcal_srv = _row_value(headers, row, "Energy Kcal (kcal) per serving")
        fat_srv = _row_value(headers, row, "Fat (g) per serving")
        satfat_srv = _row_value(headers, row, "Saturated Fat (g) per serving")
        carb_srv = _row_value(headers, row, "Carbohydrate (g) per serving")
        sugars_srv = _row_value(headers, row, "Sugars (g) per serving")
        fibre_srv = _row_value(headers, row, "Fibre (g) per serving")
        protein_srv = _row_value(headers, row, "Protein (g) per serving")
        salt_srv = _row_value(headers, row, "Salt (g) per serving")

        title = clean_lab_title(raw_title)
        serves = parse_number(servings) or 1.0
        serving_weight_g = parse_number(portion_measure)
        totals, per_serving, per_100g = _safefood_ground_truth(
            serves=serves,
            energy_kcal_srv=energy_kcal_srv,
            fat_srv=fat_srv,
            satfat_srv=satfat_srv,
            carb_srv=carb_srv,
            sugars_srv=sugars_srv,
            fibre_srv=fibre_srv,
            protein_srv=protein_srv,
            salt_srv=salt_srv,
            energy_kcal_100g=energy_kcal_100g,
            fat_100g=fat_100g,
            satfat_100g=satfat_100g,
            carb_100g=carb_100g,
            sugars_100g=sugars_100g,
            fibre_100g=fibre_100g,
            protein_100g=protein_100g,
            salt_100g=salt_100g,
        )
        recipes.append(
            RcsiLabRecipe(
                recipe_id_src=recipe_id_key,
                raw_title=str(raw_title or "").strip(),
                title=title,
                normalized_title=normalize_title(title),
                serves=serves,
                serving_weight_g=serving_weight_g,
                duration_minutes=parse_minutes(prep_time) + parse_minutes(cook_time),
                cost_category=cost_category,
                ingredient_lines=_split_lines(ingredients_raw),
                instructions=_split_lines(instructions_raw),
                ground_truth_per_serving=per_serving,
                ground_truth_per_100g=per_100g,
            )
        )
    return recipes


def lab_total_nutrients(recipe: RcsiLabRecipe) -> dict[str, float]:
    return {
        key: round(value * recipe.serves, 6)
        for key, value in recipe.ground_truth_per_serving.items()
    }


def load_safefood_web_recipes(paths: tuple[Path, ...] = DEFAULT_WEB_EXPORTS) -> list[SafefoodWebRecipe]:
    recipes: list[SafefoodWebRecipe] = []
    seen: set[str] = set()
    for path in paths:
        if not path.exists():
            continue
        for payload in json.loads(path.read_text(encoding="utf-8")):
            title = str(payload.get("name") or "").strip()
            if not title:
                continue
            key = title.lower()
            if key in seen:
                continue
            seen.add(key)
            recipes.append(
                SafefoodWebRecipe(
                    recipe_id=web_recipe_id(title),
                    title=title,
                    normalized_title=normalize_title(title),
                    url=payload.get("url"),
                    category=payload.get("category"),
                    raw=payload,
                )
            )
    return recipes


def match_lab_to_web(
    lab_recipes: list[RcsiLabRecipe],
    web_recipes: list[SafefoodWebRecipe],
    *,
    fuzzy_cutoff: float = 0.88,
) -> tuple[list[tuple[RcsiLabRecipe, SafefoodWebRecipe, str, float]], list[RcsiLabRecipe]]:
    web_by_norm = {recipe.normalized_title: recipe for recipe in web_recipes}
    web_keys = list(web_by_norm)
    matches: list[tuple[RcsiLabRecipe, SafefoodWebRecipe, str, float]] = []
    unmatched: list[RcsiLabRecipe] = []
    used_web_ids: set[str] = set()
    for lab in lab_recipes:
        web = web_by_norm.get(lab.normalized_title)
        if web and web.recipe_id not in used_web_ids:
            matches.append((lab, web, "normalized_title", 1.0))
            used_web_ids.add(web.recipe_id)
            continue
        candidates = difflib.get_close_matches(lab.normalized_title, web_keys, n=1, cutoff=fuzzy_cutoff)
        if candidates:
            web = web_by_norm[candidates[0]]
            if web.recipe_id not in used_web_ids:
                score = difflib.SequenceMatcher(None, lab.normalized_title, candidates[0]).ratio()
                matches.append((lab, web, "fuzzy_title", score))
                used_web_ids.add(web.recipe_id)
                continue
        unmatched.append(lab)
    return matches, unmatched


def rcsi_trace(lab: RcsiLabRecipe, web: SafefoodWebRecipe | None = None, match: dict[str, Any] | None = None) -> dict[str, Any]:
    trace: dict[str, Any] = {
        "nutrition_source": LAB_NUTRITION_SOURCE,
        "source": "safefood_rcsi",
        "source_label": "RCSI SafeFood lab nutrition",
        "lab_recipe_id": lab.recipe_id_src,
        "lab_title_raw": lab.raw_title,
        "lab_title_clean": lab.title,
        "ground_truth_per_serving": lab.ground_truth_per_serving,
        "ground_truth_per_100g": lab.ground_truth_per_100g,
        "serves": lab.serves,
        "serving_weight_g": lab.serving_weight_g,
        "duration_minutes": lab.duration_minutes,
        "cost_category": lab.cost_category,
    }
    if web is not None:
        trace["web_recipe_id"] = web.recipe_id
        trace["web_title"] = web.title
        trace["web_url"] = web.url
        trace["web_category"] = web.category
    if match is not None:
        trace["match"] = match
    return trace
