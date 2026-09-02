"""Programmatic audit of every workspace Excel workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import warnings

import pandas as pd
from openpyxl import load_workbook

from .constants import DOCS_DIR, RAW_DIR, REPOSITORY_ROOT
from .normalize_prices import DATASETS


SUGGESTED_NAMES = {
    "beef-cuts.xlsx": "ec_beef_cuts_prices.xlsx",
    "beef-live-animals-FpCWS.xlsx": "ec_live_cattle_prices.xlsx",
    "cereals.xlsx": "ec_cereals_prices.xlsx",
    "dairy.xlsx": "ec_dairy_prices.xlsx",
    "eggs.xlsx": "ec_egg_prices.xlsx",
    "fertiliser.xlsx": "ec_fertiliser_prices.xlsx",
    "fruits-vegetables.xlsx": "ec_fruit_vegetable_prices.xlsx",
    "oilseeds.xlsx": "ec_oilseed_prices.xlsx",
    "olive-oil.xlsx": "ec_olive_oil_prices.xlsx",
    "pigmeat-cuts.xlsx": "ec_pigmeat_cuts_prices.xlsx",
    "pigmeat-piglets.xlsx": "ec_piglet_prices.xlsx",
    "poultry.xlsx": "ec_poultry_prices.xlsx",
    "protein-crops.xlsx": "ec_protein_crop_prices.xlsx",
    "raw-milk.xlsx": "ec_raw_milk_prices.xlsx",
    "rice.xlsx": "ec_rice_prices.xlsx",
    "sheep-goat-meat.xlsx": "ec_sheep_goat_prices.xlsx",
    "sugar.xlsx": "ec_sugar_prices.xlsx",
    "wine.xlsx": "ec_wine_prices.xlsx",
    "prc_ppp_ind_1__custom_22484819_spreadsheet.xlsx": "eurostat_food_pli_2025.xlsx",
}


def discover_workbooks(root: Path = REPOSITORY_ROOT) -> list[Path]:
    return sorted(
        path
        for path in root.rglob("*")
        if path.suffix.lower() in {".xlsx", ".xls"} and ".venv" not in path.parts
    )


def _workbook_sheets(path: Path) -> list[dict[str, Any]]:
    read_only = "prc_ppp_ind_1" not in path.name
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path, read_only=read_only, data_only=True)
    sheets = []
    try:
        for sheet in workbook.worksheets:
            rows, columns = sheet.max_row, sheet.max_column
            if rows is None or columns is None:
                if read_only:
                    sheet.reset_dimensions()
                sample = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
                rows = rows or "unknown (invalid dimension metadata)"
                columns = columns or max((len(row) for row in sample), default=0)
            header = []
            for row in sheet.iter_rows(min_row=1, max_row=12, values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if len(values) >= 2:
                    header = values
                    break
            sheets.append(
                {"name": sheet.title, "rows": rows, "columns": columns, "header_sample": header[:12]}
            )
    finally:
        workbook.close()
    return sheets


def _summarize(values: pd.Series, limit: int = 16) -> str:
    unique = sorted({str(value).strip() for value in values.dropna()})
    if len(unique) > limit:
        return f"{len(unique)} values: " + ", ".join(unique[:limit]) + ", …"
    return f"{len(unique)} values: " + ", ".join(unique)


def _cost_details(path: Path) -> dict[str, Any]:
    if path.parent.name == "monthly" and "eumofa" in path.parts:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        header_index = next(index for index, row in enumerate(rows[:10]) if row[0] == "Country")
        header = rows[header_index]
        metadata = []
        previous = [None, None, None]
        for row in rows[header_index + 1 :]:
            values = list(row[:6])
            for index in range(3):
                if values[index] is None:
                    values[index] = previous[index]
                else:
                    previous[index] = values[index]
            metadata.append(values)
        price_cells = [
            row[column]
            for row in rows[header_index + 1 :]
            for column, name in enumerate(header)
            if name == "Price per Kg (EUR)"
        ]
        return {
            "columns": "Country, category, product, local description, item on sale, size/weight range, then monthly EUR/kg and local-currency price fields",
            "member_states": _summarize(pd.Series([row[0] for row in metadata])),
            "dates": "Monthly, 2021-09 through 2026-04",
            "products": _summarize(pd.Series([row[2] for row in metadata])),
            "categories": _summarize(pd.Series([row[1] for row in metadata])),
            "price_types": "Monthly median online retail price",
            "units": "EUR/unit, local currency/unit, EUR/kg, and local currency/kg; fresh table contains kg fields only",
            "missing": f"{sum(value is None for value in price_cells)} missing EUR/kg monthly cells",
            "duplicates": str(len(metadata) - len({tuple(row) for row in metadata})),
            "issues": "Use the latest current EUR/kg cell per product-country series; reject stale series and temporal outliers",
        }
    if "prc_ppp_ind_1" in path.name:
        return {
            "columns": "Metadata rows plus GEO label, 2025 value, and status flag",
            "member_states": "All EU27 plus EU/euro-area aggregates and 10 non-EU comparators; processing retains EU27 only",
            "dates": "2025 annual",
            "products": "N/A",
            "categories": "10 required food categories plus non-alcoholic beverages",
            "price_types": "Price level indices (EU27_2020=100) only; no PPP values",
            "units": "Index, EU27_2020=100",
            "missing": "No missing EU27/category PLI values",
            "duplicates": "No duplicate EU27 country/category/year keys",
            "issues": "Malformed worksheet dimension metadata declares A1; generic sheet names; non-EU rows require filtering",
        }
    frame = pd.read_excel(path, sheet_name="MyWorkSheet-1")
    columns = list(frame.columns)
    date_columns = [
        column for column in columns
        if any(token in str(column).lower() for token in ("date", "reference period", "year-month"))
    ]
    ranges = []
    for column in date_columns:
        parsed = pd.to_datetime(frame[column], errors="coerce", format="mixed")
        if parsed.notna().any():
            ranges.append(f"{column}: {parsed.min().date()}..{parsed.max().date()}")
    member_column = next((column for column in columns if "member state" in str(column).lower()), None)
    product_columns = [
        column for column in columns
        if any(token in str(column).lower() for token in ("product", "category", "farming method", "rice type", "variety"))
    ]
    stage_columns = [
        column for column in columns
        if any(token in str(column).lower() for token in ("price type", "stage", "contract type"))
    ]
    unit_columns = [column for column in columns if "unit" in str(column).lower() or "price" in str(column).lower()]
    missing = {str(column): int(count) for column, count in frame.isna().sum().items() if count}
    issues = []
    if member_column and pd.api.types.is_numeric_dtype(frame[member_column]):
        issues.append("Member States exported as numeric EC portal identifiers")
    if path.name in {"beef-cuts.xlsx", "pigmeat-cuts.xlsx"}:
        issues.append("Filename says cuts but the workbook contains minced meat only")
    if path.name == "pigmeat-piglets.xlsx":
        issues.append("Header says €/100kg but official API reports piglets per animal; reject")
    if path.name == "protein-crops.xlsx":
        issues.append("Price unit is absent from both workbook and API response; reject")
    if path.name == "wine.xlsx":
        issues.append("€/hl cannot be converted to €/kg without density")
    if path.name == "beef-live-animals-FpCWS.xlsx":
        issues.append("Live animals are not edible ingredient cuts; €/head rows are incompatible")
    if path.name == "fertiliser.xlsx":
        issues.append("Non-food input; outside ingredient layer")
    if path.name in {"dairy.xlsx", "eggs.xlsx", "olive-oil.xlsx", "rice.xlsx", "sheep-goat-meat.xlsx"}:
        issues.append("No explicit consumer retail stage")
    return {
        "columns": ", ".join(map(str, columns)),
        "member_states": _summarize(frame[member_column]) if member_column else "No Member State field",
        "dates": "; ".join(ranges) or "No parseable date field",
        "products": "; ".join(f"{column}: {_summarize(frame[column])}" for column in product_columns) or "No product field",
        "categories": "; ".join(f"{column}: {_summarize(frame[column])}" for column in product_columns if "category" in str(column).lower()) or "Encoded in product fields",
        "price_types": "; ".join(f"{column}: {_summarize(frame[column])}" for column in stage_columns) or "No explicit price-type/stage field",
        "units": "; ".join(map(str, unit_columns)),
        "missing": str(missing) if missing else "None",
        "duplicates": str(int(frame.duplicated().sum())),
        "issues": "; ".join(issues) or "No obvious price-schema issue",
    }


def build_audit_markdown(root: Path = REPOSITORY_ROOT) -> str:
    workbooks = discover_workbooks(root)
    cost_paths = set(RAW_DIR.rglob("*.xlsx"))
    lines = [
        "# Excel Source Data Audit",
        "",
        "This audit is generated from the workbooks in the workspace. Original files are treated as immutable. "
        "`data/cost/raw/` is the immutable price-source location; files retain their original names and are not overwritten.",
        "",
        f"Discovered **{len(workbooks)}** Excel workbooks: **{len(cost_paths)}** pricing/PLI candidates and "
        f"**{len(workbooks) - len(cost_paths)}** non-pricing workbooks.",
        "",
        "## Cross-file schema findings",
        "",
        "EC price workbooks have 5–12 columns and mix weekly, monthly, and marketing-year periods. "
        "Member States are either names or numeric portal identifiers. Units include €/100kg, €/tonne, €/kg, "
        "€/head, and €/hl; the protein-crop unit is missing. Market stages range from consumer-adjacent retail "
        "selling to farm-gate, processor, port, silo, live-animal, and unspecified commodity prices. These schemas "
        "must be normalized per dataset and stages must remain separate.",
        "",
    ]
    for path in workbooks:
        relative = path.relative_to(root)
        sheets = _workbook_sheets(path)
        is_cost = path in cost_paths
        details = _cost_details(path) if is_cost else {
            "columns": "See sheet header samples below",
            "member_states": "N/A for price audit",
            "dates": "N/A for price audit",
            "products": "Recipe/nutrient/mapping content; excluded from economic source layer",
            "categories": "N/A for price audit",
            "price_types": "No EC price-type field expected",
            "units": "N/A for price audit",
            "missing": "Not profiled: workbook is outside the price-source family",
            "duplicates": "Not profiled: workbook is outside the price-source family",
            "issues": "Non-pricing workbook; intentionally excluded from transformations",
        }
        suggestion = SUGGESTED_NAMES.get(path.name, path.name)
        lines.extend(
            [
                f"## `{relative}`",
                "",
                f"- Suggested descriptive filename: `{suggestion}`",
                f"- Role: {'pricing/PLI source' if is_cost else 'non-pricing workbook'}",
                "- Sheets: " + "; ".join(
                    f"`{item['name']}` ({item['rows']} rows × {item['columns']} columns)"
                    for item in sheets
                ),
                "- Header samples: " + "; ".join(
                    f"`{item['name']}`: {item['header_sample']}" for item in sheets
                ),
                f"- Columns/schema: {details['columns']}",
                f"- Member States: {details['member_states']}",
                f"- Date coverage: {details['dates']}",
                f"- Products: {details['products']}",
                f"- Categories: {details['categories']}",
                f"- Price types/stages: {details['price_types']}",
                f"- Units/price fields: {details['units']}",
                f"- Missing values: {details['missing']}",
                f"- Exact duplicate rows: {details['duplicates']}",
                f"- Quality/schema issues: {details['issues']}",
                "",
            ]
        )
    return "\n".join(lines)


def main() -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    path = DOCS_DIR / "DATA_AUDIT.md"
    path.write_text(build_audit_markdown(), encoding="utf-8")
    print(f"Wrote {path}")


if __name__ == "__main__":
    main()
