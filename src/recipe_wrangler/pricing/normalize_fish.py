"""Normalize latest EUMOFA monthly online-retail fish prices."""

from __future__ import annotations

import json
from pathlib import Path
import re
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .constants import FISH_SOURCE_DIR, PROCESSED_DIR, WINDOW_END, WINDOW_START
from .io import write_table
from .normalize_prices import NORMALIZED_COLUMNS, normalize_country
from .normalize_products import ProductMapping, map_fish_product


PRICE_MEASURE = "Price per Kg (EUR)"
SOURCE_DATASET = "eumofa_online_retail_fish"


def _header_layout(rows: list[tuple[Any, ...]]) -> tuple[int, int, int]:
    """Locate the workbook's year, month, and field-name rows."""

    header_index = next(
        (
            index
            for index, row in enumerate(rows[:10])
            if row and row[0] == "Country" and PRICE_MEASURE in row
        ),
        None,
    )
    if header_index is None:
        raise ValueError("fish_header_not_found")
    month_index = header_index - 1
    year_index = next(
        (
            index
            for index in range(header_index - 1, -1, -1)
            if any(
                isinstance(value, (int, float)) and 2020 <= int(value) <= 2100
                for value in rows[index][6:]
                if value is not None
            )
        ),
        None,
    )
    if year_index is None:
        raise ValueError("fish_year_header_not_found")
    return year_index, month_index, header_index


def _price_columns(
    rows: list[tuple[Any, ...]],
    year_index: int,
    month_index: int,
    header_index: int,
) -> list[tuple[int, pd.Timestamp]]:
    columns: list[tuple[int, pd.Timestamp]] = []
    current_year: int | None = None
    current_month: int | None = None
    for column in range(6, len(rows[header_index])):
        year = rows[year_index][column]
        month = rows[month_index][column]
        if year is not None:
            current_year = int(year)
        if month is not None:
            current_month = int(month)
        if rows[header_index][column] == PRICE_MEASURE:
            if current_year is None or current_month is None:
                raise ValueError(f"fish_date_header_missing:column_{column + 1}")
            columns.append((column, pd.Timestamp(current_year, current_month, 1)))
    if not columns:
        raise ValueError("fish_eur_per_kg_columns_not_found")
    return columns


def _temporal_outlier(latest: float, history: list[float]) -> bool:
    """Reject an implausible latest jump relative to the preceding year."""

    previous = np.asarray(history[-12:], dtype=float)
    if len(previous) < 6:
        return False
    median = float(np.median(previous))
    ratio = latest / median if median else np.inf
    return ratio < 0.2 or ratio > 5.0


def normalize_fish_prices(
    source_dir: Path = FISH_SOURCE_DIR,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Return one latest valid EUR/kg row per EUMOFA product-country series."""

    normalized: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    mappings: dict[tuple[str, str], ProductMapping] = {}
    series_count = 0
    start, end = pd.Timestamp(WINDOW_START), pd.Timestamp(WINDOW_END)
    paths = sorted(source_dir.glob("*.xlsx"))
    if not paths:
        raise ValueError(f"No EUMOFA fish workbooks found in {source_dir}")

    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            year_index, month_index, header_index = _header_layout(rows)
            price_columns = _price_columns(
                rows, year_index, month_index, header_index
            )
            previous_metadata: list[Any] = [None, None, None]
            for source_row, row in enumerate(
                rows[header_index + 1 :], start=header_index + 2
            ):
                if all(value is None for value in row):
                    continue
                series_count += 1
                metadata = list(row[:6])
                for index in range(3):
                    if metadata[index] is None:
                        metadata[index] = previous_metadata[index]
                    else:
                        previous_metadata[index] = metadata[index]
                country, category, product, description, item, size = metadata
                source_values = {
                    "country": country,
                    "category": category,
                    "product": product,
                    "product_description": description,
                    "item_on_sale": item,
                    "size_weight_range": size,
                }
                rejection_base = {
                    "source_dataset": SOURCE_DATASET,
                    "source_file": path.name,
                    "source_sheet": worksheet.title,
                    "source_row": source_row,
                    "source_values": json.dumps(
                        source_values, default=str, ensure_ascii=False
                    ),
                }
                try:
                    mapping = map_fish_product(product, item, category)
                    mappings[(mapping.raw_product, mapping.product_detail)] = mapping
                    country_code, country_name = normalize_country(country)
                    observations: list[tuple[pd.Timestamp, float]] = []
                    for column, date in price_columns:
                        value = row[column]
                        if value is None or date > end:
                            continue
                        numeric = float(value)
                        if not np.isfinite(numeric) or numeric <= 0:
                            raise ValueError(f"invalid_fish_price:{value}")
                        observations.append((date, numeric))
                    if not observations:
                        raise ValueError("fish_price_missing")
                    latest_date, latest_price = observations[-1]
                    if latest_date < start:
                        raise ValueError(f"stale_fish_price:{latest_date.date()}")
                    if _temporal_outlier(
                        latest_price, [value for _, value in observations[:-1]]
                    ):
                        raise ValueError(
                            f"latest_fish_price_temporal_outlier:{latest_price}"
                        )

                    normalized.append(
                        {
                            "source_dataset": SOURCE_DATASET,
                            "source_file": path.name,
                            "source_sheet": worksheet.title,
                            "source_country_code": country_code,
                            "source_country": country_name,
                            "product_raw": mapping.raw_product,
                            "product_normalized": mapping.canonical_product,
                            "product_detail": mapping.product_detail,
                            "food_group": mapping.food_group,
                            "price_type_raw": "Monthly median online retail price",
                            "market_stage": "retail_selling",
                            "date": latest_date.date().isoformat(),
                            "year": latest_date.year,
                            "month": latest_date.month,
                            "price_original": latest_price,
                            "unit_original": "€/kg",
                            "currency": "EUR",
                            "eur_per_kg": latest_price,
                            "pli_category": mapping.pli_category,
                            "notes": "; ".join(
                                [
                                    "EUMOFA monthly median online retail price",
                                    f"source_row={source_row}",
                                    f"product_description={description}",
                                    f"item_on_sale={item}",
                                    f"size_weight_range={size}",
                                ]
                            ),
                        }
                    )
                except (KeyError, TypeError, ValueError) as exc:
                    rejected.append({**rejection_base, "reject_reason": str(exc)})
        finally:
            workbook.close()

    normalized_frame = pd.DataFrame(normalized, columns=NORMALIZED_COLUMNS)
    rejected_frame = pd.DataFrame(rejected)
    mapping_frame = pd.DataFrame(
        [
            {"source_dataset": SOURCE_DATASET, **mapping.as_dict()}
            for mapping in mappings.values()
        ]
    ).drop_duplicates()
    assert len(normalized_frame) + len(rejected_frame) == series_count
    assert normalized_frame["eur_per_kg"].gt(0).all()
    assert normalized_frame["date"].between(WINDOW_START, WINDOW_END).all()
    assert not normalized_frame.duplicated().any()
    return normalized_frame, rejected_frame, mapping_frame


def main() -> None:
    normalized, rejected, mappings = normalize_fish_prices()
    write_table(normalized, PROCESSED_DIR / "eumofa_fish_latest_prices.csv")
    write_table(rejected, PROCESSED_DIR / "eumofa_fish_rejected.csv", parquet=False)
    write_table(mappings, PROCESSED_DIR / "eumofa_fish_product_mapping.csv", parquet=False)
    print(
        f"Normalized {len(normalized)} latest fish series; "
        f"rejected {len(rejected)}"
    )


if __name__ == "__main__":
    main()
