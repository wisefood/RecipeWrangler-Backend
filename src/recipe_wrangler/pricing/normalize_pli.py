"""Extract and validate Eurostat food Price Level Indices."""

from __future__ import annotations

import warnings
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .constants import (
    COUNTRY_NAME_TO_CODE,
    EU_COUNTRIES,
    PLI_SOURCE_DIR,
    PLI_CATEGORIES,
    PROCESSED_DIR,
)
from .io import write_table


EXPECTED_INDICATOR = "Price level indices (EU27_2020=100)"


def find_pli_workbook(source_dir: Path = PLI_SOURCE_DIR) -> Path:
    matches = sorted(source_dir.glob("*prc_ppp_ind_1*.xlsx"))
    if len(matches) != 1:
        raise ValueError(f"Expected one prc_ppp_ind_1 workbook, found {matches}")
    return matches[0]


def normalize_pli(path: Path | None = None) -> pd.DataFrame:
    """Return the required EU27 PLI categories in long form.

    The supplied Eurostat workbook declares every worksheet dimension as A1.
    Non-read-only loading is intentional because it reconstructs the real cells.
    """

    path = path or find_pli_workbook()
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Workbook contains no default style")
        workbook = load_workbook(path, read_only=False, data_only=True)

    records: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets[1:]:
            indicator = worksheet["C6"].value
            category = worksheet["C7"].value
            year = worksheet["B9"].value
            if indicator != EXPECTED_INDICATOR:
                raise ValueError(
                    f"{worksheet.title}: expected PLI indicator, found {indicator!r}"
                )
            if category not in PLI_CATEGORIES:
                continue
            for row in worksheet.iter_rows(min_row=11, values_only=True):
                country, value = row[0], row[1]
                code = COUNTRY_NAME_TO_CODE.get(country)
                if code not in EU_COUNTRIES:
                    continue
                if not isinstance(value, (int, float)):
                    raise ValueError(
                        f"{worksheet.title}: non-numeric PLI for {country}: {value!r}"
                    )
                records.append(
                    {
                        "country_code": code,
                        "country": EU_COUNTRIES[code],
                        "pli_category": category,
                        "year": int(year),
                        "pli": float(value),
                    }
                )
    finally:
        workbook.close()

    frame = pd.DataFrame.from_records(records)
    expected_rows = len(EU_COUNTRIES) * len(PLI_CATEGORIES)
    assert len(frame) == expected_rows, (len(frame), expected_rows)
    assert not frame.duplicated(["country_code", "pli_category", "year"]).any()
    assert frame["country_code"].nunique() == len(EU_COUNTRIES)
    assert frame["pli_category"].nunique() == len(PLI_CATEGORIES)
    assert frame["pli"].between(40, 250).all(), "Unexpected PPP-like values in PLI data"
    return frame.sort_values(["country_code", "pli_category"]).reset_index(drop=True)


def main() -> None:
    frame = normalize_pli()
    write_table(frame, PROCESSED_DIR / "eurostat_food_pli.csv")
    print(f"Wrote {len(frame)} validated PLI rows")


if __name__ == "__main__":
    main()
